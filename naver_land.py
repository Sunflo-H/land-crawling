import pandas as pd
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import warnings
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import json

chrome_options = Options()
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"]) # 셀레니움 로그 무시
warnings.filterwarnings("ignore", category=DeprecationWarning) # Deprecated warning 무시 

browser = webdriver.Chrome("c:/chromedriver.exe", options=chrome_options)

def 행정동리스트():
    #json 파일 열고, 데이터 변환
    file = './서울_행정동.json'
    with open(file, 'r', encoding="UTF-8") as f:
        json_data = json.load(f)
    
    행정동 = []
    for i in json_data:
        obj = {
            '시군구': i['시군구'],
            '읍면동': i['읍면동'],
            '위도': i['위도'],
            '경도': i['경도'],
            'url': 'https://new.land.naver.com/rooms?ms=%s,%s&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT' % (i['위도'],i['경도'])
            }
        행정동.append(obj)
    return 행정동


    

def 크롤링( url_index ): 
    # 브라우저를 열 url 입력
    browser.get(행정동리스트()[url_index]['url'])

    # 로딩이 끝날때까지 10초정도 기다린다.
    browser.implicitly_wait(10) 

    # 스크롤이 있는 지점을 클릭하여 스크롤이 가능하게 한다.
    scrollTarget = browser.find_element(By.CLASS_NAME, 'item_link')
    scrollTarget.click()

    # 끝까지 스크롤 하기
    beforeScroll = 0 #스크롤 전 높이

    while True:
        scrollTarget.send_keys(Keys.END)
        time.sleep(1)
        # 스크롤 했을때 scrollTop의 값이 변하는 div를 찾아 스크롤 후의 높이를 구한다.    
        afterScroll = browser.execute_script("return document.querySelector('.item_list').scrollTop") 

        if afterScroll == beforeScroll:
            break
        beforeScroll = afterScroll

    # 클릭할 아이템 리스트
    itemList = browser.find_elements(By.CLASS_NAME, 'item_link')

    # 엑셀파일 만들기
    wb = openpyxl.Workbook()

    # 엑셀 워크시트 만들기
    ws = wb.create_sheet(행정동리스트()[url_index]['시군구'])

    # 데이터 추가하기
    ws['A1'] = '시군구'
    ws['B1'] = '읍면동'
    ws['C1'] = '종류'
    ws['D1'] = '가격'
    ws['E1'] = '공급/전용면적'
    ws['F1'] = '층/총층'
    ws['G1'] = '방수/욕실수'
    ws['H1'] = '월 관리비'
    ws['I1'] = '관리비포함'
    ws['J1'] = '입주 가능일'
    ws['K1'] = '융자금'
    ws['L1'] = '사용 승인일'
    ws['M1'] = '방향'
    ws['N1'] = '주차가능여부'
    ws['O1'] = '방구조'
    ws['P1'] = '복층여부'
    ws['Q1'] = '중개사_이름'
    ws['R1'] = '중개사_사진'
    ws['S1'] = '중개사_직함'
    ws['T1'] = '중개사_직원명'
    ws['U1'] = '중개사_주소'
    ws['V1'] = '중개사_전화'

    # 22개
    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U', 'V'] 

    number = 0
    for item in itemList:
        item.click()
        time.sleep(1)

        시군구 = 행정동리스트()[url_index]['시군구']
        읍면동 = 행정동리스트()[url_index]['읍면동']
        종류 = browser.execute_script('return document.querySelector(".item_inner > .item_link > .item_title").innerText')
        가격 = browser.execute_script('return document.querySelector(".item_inner > .item_link > .price_line").innerText')
        소재지 = browser.execute_script('return document.querySelectorAll(".table_td")[0].innerText')
        매물특징 = browser.execute_script('return document.querySelectorAll(".table_td")[1].innerText')
        공급_전용면적 = browser.execute_script('return document.querySelectorAll(".table_td")[2].innerText')
        해당층_총층 = browser.execute_script('return document.querySelectorAll(".table_td")[3].innerText')
        방수_욕실수 = browser.execute_script('return document.querySelectorAll(".table_td")[4].innerText')
        월관리비 = browser.execute_script('return document.querySelectorAll(".table_td")[5].innerText')
        관리비포함 = browser.execute_script('return document.querySelectorAll(".table_td")[6].innerText')
        입주가능일 = browser.execute_script('return document.querySelectorAll(".table_td")[7].innerText')
        융자금 = browser.execute_script('return document.querySelectorAll(".table_td")[8].innerText')
        사용승인일 = browser.execute_script('return document.querySelectorAll(".table_td")[9].innerText')
        방향 = browser.execute_script('return document.querySelectorAll(".table_td")[10].innerText')
        주차가능여부 = browser.execute_script('return document.querySelectorAll(".table_td")[11].innerText')
        방구조 = browser.execute_script('return document.querySelectorAll(".table_td")[12].innerText')
        복층여부 = browser.execute_script('return document.querySelectorAll(".table_td")[13].innerText')
        건축물용도 = browser.execute_script('return document.querySelectorAll(".table_td")[14].innerText')
        매물번호 = browser.execute_script('return document.querySelectorAll(".table_td")[15].innerText')
        총주차대수 = browser.execute_script('return document.querySelectorAll(".table_td")[16].innerText')
        중개사_이름 = browser.execute_script('return document.querySelector(".info_agent_title .info_title").innerText')
        # 중개사_사진 = browser.execute_script('return document.querySelector(".info_agent_photo .info_photo_image").src')
        중개사_직함 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[0].firstElementChild.innerText')
        중개사_직원명 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[0].firstElementChild.nextElementSibling.innerText')
        중개사_주소 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[0].firstElementChild.nextElementSibling.nextElementSibling.lastElementChild.innerText')
        중개사_전화 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[1].firstElementChild.nextElementSibling.innerText')

        # 20개
        # budongsan = [종류, 가격, 공급_전용면적, 해당층_총층, 방수_욕실수, 월관리비, 관리비포함, 입주가능일, 융자금, 사용승인일, 방향, 주차가능여부, 방구조, 복층여부, 중개사_이름, 중개사_사진, 중개사_직함, 중개사_직원명, 중개사_주소, 중개사_전화]   
        # 21개
        budongsan = [시군구, 읍면동, 종류, 가격, 공급_전용면적, 해당층_총층, 방수_욕실수, 월관리비, 관리비포함, 입주가능일, 융자금, 사용승인일, 방향, 주차가능여부, 방구조, 복층여부, 중개사_이름, 중개사_직함, 중개사_직원명, 중개사_주소, 중개사_전화]
        
        for j in range(21):
            ws[abc[j] + str(number+2)] = budongsan[j]

        # 엑셀 저장하기
        wb.save('부동산_data.xlsx')
        number = number + 1
        print(number)
        print('성공')







크롤링(72)
time.sleep(2)
크롤링(74)


# url_list = [
#     {
#         'lat' : '37.5466951',
#         'lng' : '127.0914159',
#         'url' : 'https://new.land.naver.com/rooms?ms=37.5466951,127.0914159,15&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '구의동'
#     },
#     {   
#         'url' : 'https://new.land.naver.com/rooms?ms=37.546892,127.103025,16&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '광장동'
#     },
#     { 
#         'url' : 'https://new.land.naver.com/rooms?ms=37.5447823,127.0831874,15&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '구의동'
#     },
#     {
#         'url' : 'https://new.land.naver.com/rooms?ms=37.5361732,127.0761922,15&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '자양동'
#     },
#     {
#         'url' : 'https://new.land.naver.com/rooms?ms=37.5465,127.0713,16&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '화양동'
#     },
#     {
#         'url'   : 'https://new.land.naver.com/rooms?ms=37.55378,127.080499,16&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '능동'
#     },
#     {
#         'url' : 'https://new.land.naver.com/rooms?ms=37.555122,127.075837,16&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '군자동'
#     },
#     {
#         'url' : 'https://new.land.naver.com/rooms?ms=37.5631,127.0876,16&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT&ad=true',
#         '동' : '중곡동'
#     }
# ]

# print(종류) # 일반원룸
# print(가격) # 월세 500/30
# print(소재지) # 주소
# print(매물특징) # 설명
# print(공급_전용면적) #
# print(해당층_총층)
# print(방수_욕실수)
# print(월관리비)
# print(관리비포함)
# print(입주가능일)
# print(융자금)
# print(사용승인일)
# print(방향)
# print(주차가능여부)
# print(방구조)
# print(복층여부)
# print(중개사_이름)
# print(중개사_사진)
# print(중개사_직함)
# print(중개사_직원명)
# print(중개사_주소)
# print(중개사_전화)








# info1 = browser.execute_script('return document.querySelectorAll(".item_inner > .item_link > .info_area > .line")[0].innerText')
# info2 = browser.execute_script('return document.querySelectorAll(".item_inner > .item_link > .info_area > .line")[1].innerText')
# tag = browser.execute_script('return document.querySelector(".item_inner > .item_link > .tag_area").innerText')
# agentInfo1 = browser.execute_script('return document.querySelectorAll(".item_inner > .cp_area > .cp_area_inner > .agent_info")[0].innerText')
# agentInfo2 = browser.execute_script('return document.querySelectorAll(".item_inner > .cp_area > .cp_area_inner > .agent_info")[1].innerText')
# print(info1) # 필요 없음
# print(info2) # 필요없음
# print(tag) # 필요없음
# print(agentInfo1) # 필요없음
# print(agentInfo2) # 필요없음